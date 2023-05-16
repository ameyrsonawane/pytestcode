'''# Data Driven Testing
# Suppose we have some test cases which need a test data And
# we have to execute same test case multiple times with different test data this known as a data driven testing.
# We prepare the data in Excel or Database, but we use Excel to prepare the data.
# Selenium webdriver doesn't support Excel sheet,
# By openpyxl module we can work with Excel files(.xlsx) '''



# How to Read data from Excel

from openpyxl import load_workbook
wb= load_workbook(filename="C:\\Users\\ameys\\OneDrive\\Desktop\\Amey.xlsx")

ws1=wb.active              # Read Data From Sheet-1
print(ws1['A3'].value)

ws2=wb['Test Data-2']      # Read Data From Sheet-2
print(ws2['A3'].value)


row= ws1.max_row            # Number of Rows and Columns
column= ws1.max_column
print(row)
print(column)

# Read All data from excel

for i in range(1,row+1):
    for j in range(1,column+1):
        print(ws1.cell(row=i,column=j).value,end='         ')




# How to Write data from Excel

'''from openpyxl import Workbook
wb=Workbook()
ws=wb.active
#ws['A5']="Kalpak"
data=[['Name','Department','Address'],['Kalpak','Maths','Nashik'],['Pooja','Sci','Pune']]
for test in data:
    ws.append(test)
wb.save("C:\\Users\\ameys\\OneDrive\\Desktop\\Report.xlsx")'''


